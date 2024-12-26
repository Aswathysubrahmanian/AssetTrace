from django.shortcuts import render, redirect, get_object_or_404
from .models import AssetData, AssetType,AssetImage,Asset_movement_data, CustomField,Depreciation
from .forms import AssetDataForm, AssetTypeForm,AssetMovementForm
from django.urls import reverse
from .utils import generate_qr_code
from authentication.models import Branches,HeadOffice
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

#------------------------Asset listing--------------------------------------------------
@login_required()
def asset_list(request):
    user = request.user
    assets = AssetData.objects.all()
    asset_types = AssetType.objects.all()
    branches = Branches.objects.all()
    head_offices = HeadOffice.objects.all()
    search_query = request.GET.get('search',"")

    if user.is_superuser:
        
        dashboard_type = "master_admin"
        branch_id = request.GET.get('branch')
        if branch_id:
            assets = assets.filter(branch_id=branch_id)

        head_office_id = request.GET.get('asset-office')
        if head_office_id:
            assets = assets.filter(office=head_office_id)

    elif  user.head_office:
        dashboard_type = "head_office"
        # Superadmin can only see assets from their head office and its branches
        branches = branches.filter(head_office=user.head_office)
        print('branch',branches)
        assets = assets.filter(branch__in=branches).order_by('-asset_code')
        print('assets:',assets)
        asset_types = AssetType.objects.all()
        print('assets type:',asset_types)

    elif user.is_admin and user.branch:
        dashboard_type = "branch"
        # Admin can only see assets from their branch
        assets = assets.filter(branch=user.branch)
        asset_types = AssetType.objects.filter(
        Q(assets__branch=user.branch) | 
        Q(assets__isnull=True)
    )
        head_offices = head_offices.filter(branches=user.branch)

    else:
        # If user doesn't fit any of the above roles, they see nothing
        assets = AssetData.objects.none()
        asset_types = AssetType.objects.none()
    
    if search_query:
        assets = assets.filter(Q(asset_name__icontains=search_query) | Q(asset_type__name__icontains=search_query))


    asset_data = []
    for asset in assets:
        asset_url = request.build_absolute_uri(reverse('asset_link', args=[asset.id]))
        qr_code_url = generate_qr_code(asset.asset_name, asset_url)
        asset_data.append({'asset': asset, 'qr_code_url': qr_code_url})
    
    page = request.GET.get('page', 1)
    paginator = Paginator(asset_data, 10)  # Show 10 assets per page

    try:
        assets_page = paginator.page(page)
    except PageNotAnInteger:
        assets_page = paginator.page(1)
    except EmptyPage:
        assets_page = paginator.page(paginator.num_pages)
    print('asset data',assets)

    context = {
        'assets': assets_page,
        'asset_types': asset_types,
        'branches': branches,
        'head_offices': head_offices,
        'selected_branch': request.GET.get('branch'),
        'selected_head_office': request.GET.get('head_office'),
        'dashboard_type': dashboard_type,
        'search_query': search_query,
    }
    print('asset_types',asset_types)

    return render(request, 'asset/asset_list.html', context)

#-----------------------------Asset sold--------------------------------------------

from django.db.models import Q
@login_required()
def sold_asset_list(request):
    user = request.user
    search_query = request.GET.get('search', "")
    
    # Start by getting all sold assets
    sold_assets = AssetData.objects.filter(is_sold=True)

    # Filter the assets based on the type of user
    if user.is_superuser:
        # Superusers can see all sold assets
        dashboard_type = "master_admin"
        branch_id = request.GET.get('branch')
        head_office_id = request.GET.get('asset-office')

        if branch_id:
            sold_assets = sold_assets.filter(branch_id=branch_id).order_by('-sold_date')
        if head_office_id:
            sold_assets = sold_assets.filter(office=head_office_id).order_by('-sold_date')

    elif user.is_superadmin and user.head_office:
        # Superadmins can only see sold assets from their head office and its branches
        dashboard_type = "head_office"
        branches = Branches.objects.filter(head_office=user.head_office)
        sold_assets = sold_assets.filter(branch__in=branches).order_by('-sold_date')

    elif user.is_admin and user.branch:
        # Admins can only see sold assets from their branch
        dashboard_type = "branch"
        sold_assets = sold_assets.filter(branch=user.branch).order_by('-sold_date')

    # Apply search filter
    if search_query:
        sold_assets = sold_assets.filter(
            Q(asset_name__icontains=search_query) |
            Q(asset_type__name__icontains=search_query)
        )

    context = {
        'sold_assets': sold_assets,
        'search_query': search_query,
        'dashboard_type': dashboard_type,
    }

    return render(request, 'asset/sold_asset_list.html', context)

@login_required()
def mark_as_unsold(request, asset_id):
    asset = get_object_or_404(AssetData, id=asset_id)
    asset.is_sold = False
    asset.save()
    messages.success(request, f"{asset.asset_name} has been marked as unsold.")
    return redirect('sold_asset_list')


@login_required()
def sell_asset(request, pk):
    asset = get_object_or_404(AssetData, pk=pk)
    
    if request.method == "POST":
        
        asset.is_sold = True
        asset.save()
        
        messages.success(request, f'The asset "{asset.asset_name}" has been marked as sold.')
        return redirect('asset_detail', pk=asset.pk)  

    return redirect('asset_detail', pk=asset.pk)

#------------------------Asset creation--------------------------------------------------


from django.shortcuts import render, redirect
from django.urls import reverse
from .forms import AssetDataForm
from .models import CustomField, AssetImage, AssetBill
from .utils import generate_qr_code  # Assuming this is a utility to generate the QR code
import json
from django.db import transaction




from django.views.decorators.http import require_POST

@require_POST
@login_required()
def verify_asset(request, asset_id):
    asset = get_object_or_404(AssetData, id=asset_id)
    try:
        asset.mark_as_verified()
        messages.success(request, f'Asset {asset.asset_code} has been verified successfully')
        return redirect('asset_link', asset_id=asset_id)
    except Exception as e:
        messages.error(request, str(e))
        return redirect('asset_link', asset_id=asset_id)


from .forms import ImageForm
from django.forms import modelformset_factory    

      
# def asset_create(request):
#     ImageFormSet = modelformset_factory(
#         AssetImage, 
#         form=ImageForm, 
#         can_delete=True  ,
#         extra=1
#     )
#     if request.method == 'POST':
#         asset_form = AssetDataForm(request.POST, request.FILES)
#         print(asset_form)
#         if asset_form.is_valid() and image_formset.is_valid():
#             try:
#                 with transaction.atomic():
#                     asset = asset_form.save(commit=False)
#                     if asset.is_sold and not asset.sold_date:
#                         asset.sold_date = timezone.now().date()
#                     asset.save()
#                     for form in image_formset:
#                         image = form.cleaned_data.get('image')
#                         if image:
#                             AssetImage.objects.create(asset=asset, image=image)
#                     custom_fields = request.POST.get('custom_fields')
#                     if custom_fields:
#                         try:
#                             custom_fields = json.loads(custom_fields)
#                             for custom_field in custom_fields:
#                                 CustomField.objects.create(
#                                     asset=asset,
#                                     field_name=custom_field['name'],
#                                     field_value=custom_field['value']
#                                 )
#                         except (json.JSONDecodeError, KeyError) as e:
#                             messages.warning(request, f'Error processing custom fields: {str(e)}')
#                     for bill_file in request.FILES.getlist('bill_file'):
#                         AssetBill.objects.create(asset=asset, bill_file=bill_file)
#                     asset_url = request.build_absolute_uri(
#                         reverse('asset_link', args=[asset.id])
#                     )
#                     asset.qr_code_base64 = generate_qr_code(asset.asset_name, asset_url)
#                     asset.save()

                    
#                     messages.success(request, 'Asset created successfully!')
#                     return redirect('asset_list')

#             except Exception as e:
                
#                 messages.error(request, f'Error creating asset: {str(e)}')
                
#                 return render(request, 'asset/asset_form.html', {
#                     'asset_form': asset_form,
#                     'image_formset': image_formset
#                 })

#         else:
#             # Form validation failed
#             messages.error(request, 'Please correct the errors below.')
#             return render(request, 'asset/asset_form.html', {
#                 'asset_form': asset_form,
#                 'image_formset': image_formset
#             })

#     else:
#         # GET request: initialize blank forms
#         asset_form = AssetDataForm()
#         image_formset = ImageFormSet(queryset=AssetImage.objects.none())

#     # Render the form
#     return render(request, 'asset/asset_form.html', {
#         'asset_form': asset_form,
#         'image_formset': image_formset
#     })

from django.forms import modelformset_factory
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone
import json

from .models import AssetImage, CustomField, AssetBill
from .forms import AssetDataForm, ImageForm


def asset_create(request):

    if request.method == 'POST':
        asset_form = AssetDataForm(request.POST, request.FILES)
        print(asset_form)
        
        if asset_form.is_valid() :
            try:
                with transaction.atomic():
                    # Save the AssetData
                    asset = asset_form.save(commit=False)
                    if asset.is_sold and not asset.sold_date:
                        asset.sold_date = timezone.now().date()
                    asset.save()

                    # Save associated images
                    for image in request.FILES.getlist('image'):
                        
                            AssetImage.objects.create(asset=asset, image=image)

                    # Handle custom fields
                    custom_fields = request.POST.get('custom_fields')
                    if custom_fields:
                        try:
                            custom_fields = json.loads(custom_fields)
                            for custom_field in custom_fields:
                                CustomField.objects.create(
                                    asset=asset,
                                    field_name=custom_field['name'],
                                    field_value=custom_field['value']
                                )
                        except (json.JSONDecodeError, KeyError) as e:
                            messages.warning(request, f'Error processing custom fields: {str(e)}')

                    # Save bill files
                    for bill_file in request.FILES.getlist('bill_file'):
                        AssetBill.objects.create(asset=asset, bill_file=bill_file)

                    # Generate QR code and save asset
                    asset_url = request.build_absolute_uri(reverse('asset_link', args=[asset.id]))
                    asset.qr_code_base64 = generate_qr_code(asset.asset_name, asset_url)
                    asset.save()

                    messages.success(request, 'Asset created successfully!')
                    return redirect('asset_list')

            except Exception as e:
                messages.error(request, f'Error creating asset: {str(e)}')

        else:
            # Form validation failed
            messages.error(request, "{}".format(asset_form.errors))

    else:
        # Initialize forms for GET request
        asset_form = AssetDataForm()

    return render(request, 'asset/asset_form.html', {
        'asset_form': asset_form,
    })
    
    
from .models import AssetBill
    
from django.http import JsonResponse
from .forms import AssetTypeForm
@login_required()
def create_asset_type(request):
    if request.method == 'POST':
        form = AssetTypeForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            return redirect('asset_list')
        else:
            # If the form is invalid and an AJAX request, return errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = AssetTypeForm()
    
    return render(request, 'asset/asset_type.html', {'form': form})


# def get_branches(request):
#     office_id = request.GET.get('office_id')
#     print(office_id)
#     branches = Branches.objects.filter(head_office_id=office_id).values('id', 'branch_name')
#     print(branches)
#     return JsonResponse(list(branches), safe=False)

def get_branches(request):
    office_id = request.GET.get('office_id')
    if office_id:
        branches = Branches.objects.filter(head_office_id=office_id).values('id', 'branch_name')
        return JsonResponse(list(branches), safe=False)
    return JsonResponse([], safe=False)

def get_branch_prefix(request, branch_id):
    branch = Branches.objects.get(id=branch_id)
    return JsonResponse({'prefix': branch.prefix})

def generate_asset_code(request, branch_id):
    branch = Branches.objects.get(id=branch_id)
    temp_asset = AssetData(branch=branch)
    asset_code = temp_asset.generate_asset_code()
    return JsonResponse({'asset_code': asset_code})


#------------------------Asset Details--------------------------------------------------
@login_required()
def asset_detail(request, pk):
    asset = get_object_or_404(AssetData, pk=pk)
        
    asset_url = request.build_absolute_uri(reverse('asset_link', args=[asset.id]))
    qr_code_base64 = generate_qr_code(asset.asset_name, asset_url)
    
    context={
        'custom_fields': asset.custom_fields if asset.custom_fields else {},
        'asset': asset,
        'qr_code_base64': qr_code_base64,
    }
    return render(request, 'asset/asset_detail.html', context)




#------------------------Asset edit--------------------------------------------------

@login_required()
def asset_type_edit(request, pk):
    asset = get_object_or_404(AssetType, pk=pk)

    if request.method == 'POST':
        asset_type_form = AssetTypeForm(request.POST, instance=asset, prefix='asset_type')

        if 'asset_type_submit' in request.POST and asset_type_form.is_valid():
            asset_type_form.save()
            return redirect('asset_list')
    else:
        asset_type_form = AssetTypeForm(instance=asset, prefix='asset_type')

    return render(request, 'asset/edit_assettype.html', {'asset_type_form': asset_type_form})


@login_required()
def asset_edit(request, pk):
    asset = get_object_or_404(AssetData, pk=pk)
    asset_images = AssetImage.objects.filter(asset=asset)  
    print('asset_images', asset_images)
    print('asset', asset.office)
    
    ImageFormSet = modelformset_factory(AssetImage, form=ImageForm, extra=1, can_delete=True)
    
    if request.method == 'POST':
        asset_form = AssetDataForm(request.POST, request.FILES, instance=asset, prefix='asset')
        image_formset = ImageFormSet(request.POST, request.FILES, queryset=asset_images, prefix='images')

        if 'asset_submit' in request.POST and asset_form.is_valid() and image_formset.is_valid():
            asset = asset_form.save(commit=False)
            asset_url = request.build_absolute_uri(reverse('asset_link', args=[asset.id]))
            asset.qr_code_base64 = generate_qr_code(asset.asset_name, asset_url)
            asset.save()
            
            # Save all images in the formset
            for form in image_formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    image = form.save(commit=False)
                    image.asset = asset
                    image.save()

            # Handle custom fields
            custom_fields = request.POST.get('custom_fields')
            if custom_fields:
                custom_fields = json.loads(custom_fields)
                for custom_field in custom_fields:
                    CustomField.objects.update_or_create(
                        asset=asset,
                        field_name=custom_field['name'],
                        defaults={'field_value': custom_field['value']}
                    )

            messages.success(request, 'Asset and images updated successfully.')
            return redirect('asset_list')
        else:
            # Log errors for debugging
            print('Asset Form Errors:', asset_form.errors)
            print('Image Formset Errors:', image_formset.errors)
            messages.error(request, 'There was an error updating the asset or images. Please check the form.')
    else:
        asset_form = AssetDataForm(instance=asset, prefix='asset')
        image_formset = ImageFormSet(queryset=asset_images, prefix='images')  # Use formset to handle multiple images

    custom_fields = CustomField.objects.filter(asset=asset)
    return render(request, 'asset/asset_edit.html', {
        'asset_form': asset_form,
        'custom_fields': custom_fields,
        'asset': asset,  
        'image_formset': image_formset  # Pass the formset to the template
    })
    
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import AssetImage

@csrf_exempt
def delete_image(request, image_id):
    if request.method == 'POST':
        try:
            image = AssetImage.objects.get(id=image_id)
            image.delete()
            return JsonResponse({'success': True})
        except AssetImage.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Image not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

#------------------------Asset delete--------------------------------------------------

@login_required()
def asset_delete(request, pk):
    asset = get_object_or_404(AssetData, pk=pk)
    
    asset.delete()
    return redirect('asset_list')
@login_required()   
def asset_type_delete(request, pk):
    asset = get_object_or_404(AssetType, pk=pk)
    
    asset.delete()
    return redirect('asset_list')


#------------------------Asset QR page--------------------------------------------------


def asset_link(request, asset_id):
    
    asset = get_object_or_404(AssetData, pk=asset_id)
    depreciations = Depreciation.objects.filter(asset=asset)
    context={
        'custom_fields': asset.custom_fields if asset.custom_fields else {},
        'asset': asset,
        'depreciations':depreciations
        
    }
    if request.user.is_authenticated:
        return render(request, 'asset/asset_full_details.html', {'asset': asset})
    return render(request, 'asset/asset_link.html', context)


from django.http import HttpResponse, FileResponse
from django.conf import settings
import os

import os
from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
import logging

# Get an instance of a logger
logger = logging.getLogger(__name__)

# def download_qr_code(request, asset_id):
#     # Fetch the asset object from the database
#     asset = get_object_or_404(AssetData, id=asset_id)
    
#     # Construct the file path where the QR code is stored
#     qr_code_path = os.path.join(settings.MEDIA_ROOT, 'qr_codes', f'{asset.asset_name}.png')
    
#     # Check if the QR code file exists
#     if not os.path.exists(qr_code_path):
#         # Log the event and return a 404 response
#         print(f"QR code file not found for asset: {asset.asset_name}")
#         return HttpResponse("QR code not found", status=404)
    
#     try:
#         # Serve the QR code as a downloadable file using FileResponse
#         response = FileResponse(open(qr_code_path, 'rb'), content_type='image/png')
#         response['Content-Disposition'] = f'attachment; filename="{asset.asset_name}_qr.png"'
#         return response
    
#     except Exception as e:
#         # Log the error with details for debugging
#         print(f"Error downloading QR code for asset {asset.asset_name}: {str(e)}")
#         return HttpResponse("Error downloading QR code", status=500)

from django.conf import settings
from django.http import FileResponse, HttpResponse, Http404
from django.shortcuts import get_object_or_404
from PIL import Image, ImageDraw, ImageFont
import os

from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from PIL import Image, ImageDraw, ImageFont
import os

from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from PIL import Image, ImageDraw, ImageFont
import os
@login_required()
def download_qr_code(request, asset_id):
    # Fetch the asset object from the database
    asset = get_object_or_404(AssetData, id=asset_id)

    # Fetch the head office object (assuming you have a head office model)
    head_office = get_object_or_404(HeadOffice)  # Adjust this according to your model

    # Construct the file path where the QR code is stored
    qr_code_path = os.path.join(settings.MEDIA_ROOT, 'qr_codes', f'{asset.asset_name}.png')

    # Check if the QR code file exists
    if not os.path.exists(qr_code_path):
        print(f"QR code file not found for asset: {asset.asset_name}")
        return HttpResponse("QR code not found", status=404)

    try:
        # Load the QR code image
        qr_code_img = Image.open(qr_code_path)

        # Load the head office logo
        head_office_logo_path = os.path.join(settings.MEDIA_ROOT, 'offiec_logo', head_office.logo_image.name)
        if os.path.exists(head_office_logo_path):
            logo = Image.open(head_office_logo_path)

            # Resize logo to fit on the QR code
            logo.thumbnail((50, 50), Image.ANTIALIAS)  # Adjust size as needed

            # Paste the logo onto the QR code (centered)
            qr_code_img.paste(logo, ((qr_code_img.width - logo.width) // 2, 
                                     (qr_code_img.height - logo.height) // 2), logo)

        # Prepare to add the asset code text below the QR code
        draw = ImageDraw.Draw(qr_code_img)
        font_path = os.path.join(settings.MEDIA_ROOT, 'fonts', 'arial.ttf')  # Make sure the font file exists

        try:
            font = ImageFont.truetype(font_path, 20)  # Adjust font size as needed
        except IOError:
            font = ImageFont.load_default()  # Fallback if the font isn't found

        # Position the asset code at the bottom of the image
        asset_code_text = f'Asset Code: {asset.asset_code}'
        
        # Calculate the size of the text using the textbbox method
        text_bbox = draw.textbbox((0, 0), asset_code_text, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]

        # Position the text at the bottom, centered
        text_position = ((qr_code_img.width - text_width) // 2, qr_code_img.height - text_height - 10)

        # Draw the asset code onto the image
        draw.text(text_position, asset_code_text, font=font, fill='black')

        # Save the modified image to a temporary path
        output_path = os.path.join(settings.MEDIA_ROOT, 'qr_codes', f'{asset.asset_name}_with_logo.png')
        qr_code_img.save(output_path)

        # Serve the QR code with the logo and asset code as a downloadable file
        response = FileResponse(open(output_path, 'rb'), content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="{asset.asset_name}_qr_with_logo.png"'
        return response

    except Exception as e:
        print(f"Error downloading QR code for asset {asset.asset_name}: {str(e)}")
        return HttpResponse("Error downloading QR code", status=500)

#------------------------Asset movement--------------------------------------------------


from django.db.models import Q
@login_required()
def asset_movement_page(request):
    if request.method == 'POST':
        form = AssetMovementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('asset_movement_list')
        else:
            print(form.errors)
    else:
        form = AssetMovementForm()
    
    user = request.user
    if user.is_superuser:
        movements = Asset_movement_data.objects.all()
    elif user.is_superadmin and user.head_office:
        branches = Branches.objects.filter(head_office=user.head_office)
        movements = Asset_movement_data.objects.filter(
            Q(from_branch__in=branches) | Q(to_branch__in=branches)
        )
    elif user.is_admin and user.branch:
        movements = Asset_movement_data.objects.filter(
            Q(from_branch=user.branch) | Q(to_branch=user.branch)
        )
    else:
        movements = Asset_movement_data.objects.none()
    return render(request, 'asset/asset_movement_list.html', {'form': form, 'movements': movements})

# def get_asset_branch(request):
#     asset_id = request.GET.get('id_asset')
#     if asset_id:
#         try:
#             asset = AssetData.objects.get(id=asset_id)
#             return JsonResponse({'branch_id': asset.branch.id, 'branch_name': asset.branch.branch_name})
#         except AssetData.DoesNotExist:
#             return JsonResponse({'branch_id': None, 'branch_name': None})
#     return JsonResponse({'branch_id': None, 'branch_name': None})

# def get_asset_branch(request):
#     asset_id = request.GET.get('id_asset')
#     if not asset_id:
#         return JsonResponse({'branch_id': None, 'branch_name': None, 'error': 'Asset ID not provided'}, status=400)

#     try:
#         asset = AssetData.objects.get(id=asset_id)
#         branch_id = asset.branch.id if asset.branch else None
#         branch_name = asset.branch.branch_name if asset.branch else None
#         return JsonResponse({'branch_id': branch_id, 'branch_name': branch_name})
#     except AssetData.DoesNotExist:
#         return JsonResponse({'branch_id': None, 'branch_name': None, 'error': 'Asset not found'}, status=404)
#     except Exception as e:
#         return JsonResponse({'branch_id': None, 'branch_name': None, 'error': str(e)}, status=500)
    
def get_asset_branch(request):
    asset_id = request.GET.get('id_asset')
    if asset_id:
        asset = AssetData.objects.get(id=asset_id)
        current_branch = Asset_movement_data.get_current_branch(asset)
        return JsonResponse({
            'branch_id': current_branch.id,
            'branch_name': current_branch.branch_name
        })
    return JsonResponse({})  
    
 #------------------------Asset movement edit--------------------------------------------------
   
@login_required()   
def edit_asset_movement(request, movement_id):
    # Get the specific movement entry by ID or return a 404 error
    movement = get_object_or_404(Asset_movement_data, id=movement_id)
    print(movement)

    if request.method == 'POST':
        # Bind the form to the POST data and the current instance
        form = AssetMovementForm(request.POST, instance=movement)
        print(form)
        
        if form.is_valid():
            form.save()  # Save the changes
            return redirect('asset_movement_list')  # Redirect to the asset movement list
    else:
        # Instantiate the form with the current movement instance for GET requests
        form = AssetMovementForm(instance=movement)
    
    # Render the edit template with the form and movement instance
    return render(request, 'asset/edit_asset_movement.html', {'form': form, 'movement': movement})



#---------------------------------asset shifting filtering---------------------------------------

def asset_filter_by_branches(request, branch_id):
    branch = Branches.objects.get(id=branch_id)
    departments=Departments.objects.filter(branch=branch)
    moved_out_assets = Asset_movement_data.objects.filter(from_branch=branch).values_list('asset_id', flat=True)
    assets = AssetData.objects.filter(branch=branch).exclude(id__in=moved_out_assets)
    moved_assets = Asset_movement_data.objects.filter(from_branch=branch).order_by('-movement_date')
    received_assets = Asset_movement_data.objects.filter(to_branch=branch).order_by('-movement_date')
    
    context = {
        'branch': branch,
        'assets': assets,
        'moved_assets': moved_assets,
        'received_assets': received_assets,
        'departments':departments
    }
    return render(request, 'asset/assets_by_branch.html', context)




@login_required()
def asset_movement_list(request):
    
    user = request.user
    if user.is_superuser:
        movements = Asset_movement_data.objects.all()
    elif user.is_superadmin and user.head_office:
        branches = Branches.objects.filter(head_office=user.head_office)
        movements = Asset_movement_data.objects.filter(
            Q(from_branch__in=branches) | Q(to_branch__in=branches)
        )
    else:
        movements = Asset_movement_data.objects.none()
    
    form = AssetMovementForm()  # Assuming you have this form
    
    return render(request, 'authentication/asset_movement_list.html', {
        'movements': movements,
        'form': form
    })
    
    
from django.contrib import messages
    



# def depreciation_list(request):
#     user = request.user
#     depreciations = Depreciation.objects.all()
#     branches = Branches.objects.all()
#     head_offices = HeadOffice.objects.all()

#     if user.is_superuser:
#         # Superuser can see all assets but can still filter
#         branch_id = request.GET.get('branch')
#         if branch_id:
#             depreciations = depreciations.filter(asset__branch_id=branch_id)

#         head_office_id = request.GET.get('head_office')
#         if head_office_id:
#             depreciations = depreciations.filter(asset__branch__head_office_id=head_office_id)

#     elif user.is_superadmin and user.head_office:
#         # Superadmin can only see assets from their head office and its branches
#         branches = branches.filter(head_office=user.head_office)
#         depreciations = depreciations.filter(asset__branch__in=branches)

#     elif user.is_admin and user.branch:
#         # Admin can only see assets from their branch
#         depreciations = depreciations.filter(asset__branch=user.branch)
#         head_offices = head_offices.filter(branches=user.branch)

#     else:
#         # If user doesn't fit any of the above roles, they see nothing
#         depreciations = Depreciation.objects.none()

#     context = {
#         'depreciations': depreciations,
#         'branches': branches,
#         'head_offices': head_offices,
#     }

#     return render(request, 'asset/depreciation_calculated.html', context)






#------------------------Asset Depreciation calculation--------------------------------------------------



from .forms import DepreciationForm, CustomDepreciationForm
import json
from decimal import Decimal, InvalidOperation

@login_required()
def create_depreciation(request):
    asset_id = request.GET.get('asset')
    asset = get_object_or_404(AssetData, id=asset_id)

    if request.method == 'POST':
        form = DepreciationForm(request.POST, asset=asset)
        if form.is_valid():
            depreciation = form.save(commit=False)
            depreciation.asset = asset
            depreciation.purchase_value = Decimal(str(asset.price))
            depreciation.save()
            messages.success(request, f'Depreciation for asset {asset.asset_name} created successfully.')
            return redirect('depreciation_detail', pk=depreciation.pk)
        else:
            messages.error(request, f'Error creating depreciation: {form.errors}')
    else:
        form = DepreciationForm(asset=asset)

    return render(request, 'asset/create_depreciation.html', {'form': form, 'asset': asset})




@login_required()
def create_custom_depreciation(request):
    asset_id = request.GET.get('asset')
    asset = get_object_or_404(AssetData, id=asset_id)

    if request.method == 'POST':
        form = CustomDepreciationForm(request.POST, asset=asset)
        if form.is_valid():
            depreciation = form.save(commit=False)
            depreciation.asset = asset
            depreciation.purchase_value = asset.price
            depreciation.depreciation_method = 'other'
            depreciation.custom_percentages = form.cleaned_data['custom_percentages']
            depreciation.save()
            messages.success(request, f'Custom depreciation for asset {asset.asset_name} created successfully.')
            return redirect('depreciation_detail', pk=depreciation.pk)
        else:
            messages.error(request, f'Error creating custom depreciation: {form.errors}')
    else:
        form = CustomDepreciationForm(asset=asset)

    return render(request, 'asset/create_custom_depreciation.html', {'form': form, 'asset': asset})



@login_required()
def depreciation_detail(request, pk):
    depreciation = get_object_or_404(Depreciation, pk=pk)
    depreciation_data = depreciation.calculate_depreciation()

    accumulated_depreciation = 0
    book_value = float(depreciation.purchase_value)
    schedule = []

    for year, amount in depreciation_data:
        accumulated_depreciation += amount
        book_value -= amount
        schedule.append({
            'year': year,
            'amount': amount,
            'accumulated_depreciation': accumulated_depreciation,
            'book_value': book_value
        })

    context = {
        'asset': depreciation.asset,
        'depreciation': depreciation,
        'depreciation_schedule': schedule
    }
    return render(request, 'asset/depreciation_detail.html', context)    
    
from django.shortcuts import render
from .models import AssetData, Depreciation
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q

@login_required
def asset_depreciation_list(request):
    user = request.user
    assets = AssetData.objects.all()
    search_query = request.GET.get('search', "")

    if user.is_superuser:
        dashboard_type = "master_admin"
        branch_id = request.GET.get('branch')
        if branch_id:
            assets = assets.filter(branch_id=branch_id)

        head_office_id = request.GET.get('asset-office')
        if head_office_id:
            assets = assets.filter(office=head_office_id)

    elif user.is_superadmin and user.head_office:
        dashboard_type = "head_office"
        branches = Branches.objects.filter(head_office=user.head_office)
        assets = assets.filter(branch__in=branches)

    elif user.is_admin and user.branch:
        dashboard_type = "branch"
        assets = assets.filter(branch=user.branch)

    else:
        assets = AssetData.objects.none()
    
    if search_query:
        assets = assets.filter(Q(asset_name__icontains=search_query) | Q(asset_type__name__icontains=search_query))

    # asset_depreciation_data = []
    # for asset in assets:
    #     depreciation = Depreciation.objects.filter(asset=asset).first()
    #     if depreciation:
    #         depreciation_data = depreciation.calculate_depreciation()
    #         latest_depreciation = depreciation_data[-1] if depreciation_data else None
    #         asset_depreciation_data.append({
    #             'asset': asset,
    #             'depreciation': depreciation,
    #             'latest_depreciation': latest_depreciation
    #         })
    #     else:
    #         asset_depreciation_data.append({
    #             'asset': asset,
    #             'depreciation': None,
    #             'latest_depreciation': None
    #         })
    asset_depreciation_data = []
    for asset in assets:
            try:
                depreciation = Depreciation.objects.filter(asset=asset).first()
                if depreciation:
                    try:
                        depreciation_data = depreciation.calculate_depreciation()
                        latest_depreciation = depreciation_data[-1] if depreciation_data else None
                        asset_depreciation_data.append({
                            'asset': asset,
                            'depreciation': depreciation,
                            'latest_depreciation': latest_depreciation
                        })
                    except (InvalidOperation, TypeError, ValueError) as e:
                        # Log the error and continue with next asset
                        asset_depreciation_data.append({
                            'asset': asset,
                            'depreciation': None,
                            'latest_depreciation': None,
                            'error': str(e)
                        })
                else:
                    asset_depreciation_data.append({
                        'asset': asset,
                        'depreciation': None,
                        'latest_depreciation': None
                    })
            except Exception as e:
                # Handle individual asset errors
                continue

    page = request.GET.get('page', 1)
    paginator = Paginator(asset_depreciation_data, 10)  # Show 10 assets per page

    try:
        assets_page = paginator.page(page)
    except PageNotAnInteger:
        assets_page = paginator.page(1)
    except EmptyPage:
        assets_page = paginator.page(paginator.num_pages)

    context = {
        'assets': assets_page,
        'dashboard_type': dashboard_type,
        'search_query': search_query,
    }

    return render(request, 'asset/asset_depreciation_list.html', context)


@login_required()
def delete_depreciation(request, pk):
    depreciation = get_object_or_404(Depreciation, pk=pk)
    asset_name = depreciation.asset.asset_name
    
    depreciation.delete()
    messages.success(request, f'Depreciation for asset {asset_name} deleted successfully.')
    return redirect('asset_depreciation_list')


@login_required()
def edit_standard_depreciation(request, pk):
    depreciation = get_object_or_404(Depreciation, pk=pk)
    
    if request.method == 'POST':
        form = DepreciationForm(request.POST, instance=depreciation)
        if form.is_valid():
            form.save()
            messages.success(request, 'Standard depreciation updated successfully.')
            return redirect('depreciation_detail', pk=depreciation.pk)
    else:
        form = DepreciationForm(instance=depreciation)
    
    return render(request, 'asset/edit_standard_depreciation.html', {'form': form, 'depreciation': depreciation})



@login_required()
def edit_custom_depreciation(request, pk):
    depreciation = get_object_or_404(Depreciation, pk=pk)
    
    if request.method == 'POST':
        form = CustomDepreciationForm(request.POST, instance=depreciation)
        if form.is_valid():
            form.save()
            messages.success(request, 'Custom depreciation updated successfully.')
            return redirect('depreciation_detail', pk=depreciation.pk)
    else:
        form = CustomDepreciationForm(instance=depreciation)
    
    return render(request, 'asset/edit_custom_depreciation.html', {'form': form, 'depreciation': depreciation})


#------------------------Asset Data in excel/pdf--------------------------------------------------






from django.http import HttpResponse
import xlsxwriter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from datetime import datetime
import io
import os
from django.conf import settings
from .models import AssetData, CustomField, HeadOffice
@login_required()
def export_excel(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # Add headers (including space for custom fields)
    headers = [
        'Asset Code', 
        'Asset Type', 
        'Office',
        'Branch',
        'Department',
        'Asset Name',
        'Model',
        'Serial Number',
        'Purchase Date',
        'Warranty Info',
        'Price',
        'Status'
    ]

    # Get all unique custom field names
    custom_field_names = CustomField.objects.values_list('field_name', flat=True).distinct()
    headers.extend(custom_field_names)

    # Header format
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4B5563',
        'color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'text_wrap': True
    })

    # Data format
    data_format = workbook.add_format({
        'align': 'left',
        'valign': 'top',
        'border': 1,
        'text_wrap': True
    })

    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Set column widths
    worksheet.set_column('A:Z', 20)  # Set width for all columns (extended to Z)

    # Get the data
    assets = AssetData.objects.all().select_related(
        'asset_type', 
        'office', 
        'branch', 
        'department'
    ).prefetch_related('custom_fields')
    
    # Write data rows
    for row, asset in enumerate(assets, 1):
        data = [
            asset.asset_code,
            asset.asset_type.name,
            asset.office.name,
            asset.branch.branch_name,
            asset.department.dpt_name,
            asset.asset_name,
            asset.model or 'N/A',
            asset.serial_number or 'N/A',
            asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else 'N/A',
            asset.warranty_info or 'N/A',
            asset.price,
            'Sold' if asset.is_sold else 'Available'
        ]
        
        # Add custom field values
        custom_fields = {cf.field_name: cf.field_value for cf in asset.custom_fields.all()}
        for field_name in custom_field_names:
            data.append(custom_fields.get(field_name, 'N/A'))
        
        for col, value in enumerate(data):
            worksheet.write(row, col, value, data_format)

    workbook.close()
    
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asset_report.xlsx'
    
    return response

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.units import inch
from django.http import HttpResponse
from datetime import datetime
import os
from django.conf import settings

def get_custom_field_headers(max_fields=15):
    custom_fields = CustomField.objects.values_list('field_name', flat=True).distinct()
    return list(custom_fields)[:max_fields]

def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="asset_report.pdf"'

    # Get HeadOffice details
    head_office = HeadOffice.objects.first()

    # Create custom styles
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_RIGHT,
    )

    address_style = ParagraphStyle(
        'AddressStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_RIGHT
    )

    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.grey
    )

    title_style = ParagraphStyle(
    'CustomTitle',
    parent=styles['Heading1'],
    fontSize=16,  # Reduced from 18
    spaceAfter=20,  # Reduced from 30
    spaceBefore=40,  # Reduced from 60
    alignment=TA_CENTER
)
    
    italic_style = ParagraphStyle(
        'ItalicStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.black,
        italic=True,
        spaceBefore=20
    )
# Enhanced table style for better readability
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ])

    class PdfReport(SimpleDocTemplate):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.head_office = kwargs.get('head_office')
            self.generated_by = kwargs.get('generated_by', 'Unknown User')
            self.generated_on = kwargs.get('generated_on', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            self.header_style = header_style
            self.footer_style = footer_style
            self.address_style = address_style

        def header(self, canvas, doc):
            canvas.saveState()
            top_margin = doc.height + doc.topMargin

            # Adjust logo position
            logo_y_position = doc.height + doc.topMargin - 15  # Adjusted from -25
            if self.head_office.logo_image:
                logo_path = os.path.join(settings.MEDIA_ROOT, str(self.head_office.logo_image))
                canvas.drawImage(logo_path, doc.leftMargin, logo_y_position, 
                            width=2*inch, height=0.8*inch, preserveAspectRatio=True)

            # Header text placement
            header_text_y_position = logo_y_position - 15  # Adjusted from -25
            address_text = f"""
            <para alignment="right">
            <b>{self.head_office.name}</b><br/>
            {self.head_office.address}<br/>
            Contact: {self.head_office.contact_number}
            </para>
            """
            
            p_address = Paragraph(address_text, self.address_style)
            p_address.wrapOn(canvas, doc.width/2, doc.topMargin)  # Adjusted width
            p_address.drawOn(canvas, doc.width/2 + doc.leftMargin, header_text_y_position)

            # Adjust separator line position
            canvas.line(doc.leftMargin, top_margin - 55,  # Adjusted from -75
                    doc.width + doc.leftMargin, top_margin - 55)

            # Adjust generation info position
            generated_info_y_position = top_margin - 70  # Adjusted from -90
            generated_info_left = f"""
            <para alignment="left">
            Report Generated on: {self.generated_on}
            </para>
            """
            p_info_left = Paragraph(generated_info_left, self.header_style)
            p_info_left.wrapOn(canvas, doc.width/2 - doc.leftMargin, doc.topMargin)
            p_info_left.drawOn(canvas, doc.leftMargin, generated_info_y_position)

            generated_info_right = f"""
            <para alignment="right">
            Report Generated by: {self.generated_by}
            </para>
            """
            p_info_right = Paragraph(generated_info_right, self.header_style)
            p_info_right.wrapOn(canvas, doc.width/2 - doc.leftMargin, doc.topMargin)
            p_info_right.drawOn(canvas, doc.width/2 + doc.leftMargin, generated_info_y_position)

            canvas.restoreState()
            
        def footer(self, canvas, doc):
            canvas.saveState()
            footer_margin = doc.bottomMargin - 10

            canvas.line(doc.leftMargin, footer_margin + 15, 
                       doc.width + doc.leftMargin, footer_margin + 15)

            footer_text = "@ Gaude AI Solution - Powered by Asset Trace"
            p = Paragraph(footer_text, self.footer_style)
            p.wrapOn(canvas, doc.width, doc.bottomMargin)
            p.drawOn(canvas, doc.leftMargin, footer_margin)

            page_num = canvas.getPageNumber()
            canvas.drawRightString(doc.width + doc.leftMargin, footer_margin, 
                                 f"Page {page_num}")

            canvas.restoreState()

    def myFirstPage(canvas, doc):
        doc.header(canvas, doc)
        doc.footer(canvas, doc)

    def myLaterPages(canvas, doc):
        doc.header(canvas, doc)
        doc.footer(canvas, doc)

    # Enhanced wrap_text function
    def wrap_text(text, max_length=150):
        try:
            text = str(text)
            if text is None or text.lower() == 'none' or text.strip() == '':
                return Paragraph('N/A', 
                    ParagraphStyle(
                        'Normal',
                        parent=styles['Normal'],
                        fontSize=8,
                        leading=10,
                        wordWrap='CJK'
                    ))
            
            if len(text) > max_length:
                text = text[:max_length-3] + '...'
            
            return Paragraph(
                text,
                ParagraphStyle(
                    'Normal',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10,
                    wordWrap='CJK'
                )
            )
        except Exception:
            return Paragraph('Error', styles['Normal'])
    elements = []
    elements.append(Paragraph('Assets Report', title_style))

    # Get all assets with their related data
    assets = AssetData.objects.all().select_related('asset_type', 'department').prefetch_related('custom_fields')
    
    # Get custom field headers
    custom_field_headers = get_custom_field_headers()
    
    # Define base headers and their relative widths
    base_headers_and_widths = [
        ('Asset Code',1.2),
        ('Asset Name', 1.5),
        ('Model', 1.2),
        ('Serial Number', 1.2),
        ('Purchase Date', 0.9),
        ('Warranty', 0.9),
        ('Price', 0.8),
        ('Status', 0.8),
    ]

    # Add custom field headers with default width
    for custom_header in custom_field_headers:
        base_headers_and_widths.append((custom_header, 1.2))

    headers = [h[0] for h in base_headers_and_widths]
    col_widths = [w[1] * inch for w in base_headers_and_widths]

    # Prepare table data
    data = [[wrap_text(header) for header in headers]]

    for asset in assets:
        # Get custom fields for this asset
        custom_fields = {cf.field_name: cf.field_value for cf in asset.custom_fields.all()}
        
        # Add base fields
        row = [
            wrap_text(asset.asset_code),
            wrap_text(asset.asset_name),
            wrap_text(asset.model),
            wrap_text(asset.serial_number),
            wrap_text(asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else 'N/A'),
            wrap_text(asset.warranty_info),
            wrap_text(asset.price),
            wrap_text('Available' if not asset.is_sold else 'Sold'),
        ]

        # Add custom fields
        for custom_header in custom_field_headers:
            value = custom_fields.get(custom_header, 'N/A')
            row.append(wrap_text(value))

        data.append(row)

    # Adjust page size based on number of columns
    total_columns = len(headers)
    custom_page_width = max(20, (total_columns * 1.2)) * inch
    custom_page_height = 13 * inch
    custom_page_size = (custom_page_width, custom_page_height)

    # Create document
    doc = PdfReport(
        response,
        pagesize=landscape(custom_page_size),
        rightMargin=10,
        leftMargin=10,
        topMargin=90,
        bottomMargin=30,
        head_office=head_office,
        generated_by=request.user.username,
        generated_on=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )

    # Create table
    table = Table(data, 
                 colWidths=col_widths, 
                 repeatRows=1,
                 )
    table.setStyle(table_style)
    
    elements.append(table)

    # Add footer note
    footer_note = Paragraph(
        "(This document is computer generated and does not require the Registrar's signature or the Company's stamp in order to be considered valid.)",
        italic_style
    )
    elements.append(footer_note)

    # Build the PDF
    doc.build(elements, onFirstPage=myFirstPage, onLaterPages=myLaterPages)
    
    return response

from django.views.generic import ListView
from django.db.models import Sum, Q
from django.shortcuts import render
from decimal import Decimal
from datetime import datetime
from authentication.models import Branches,Departments
from time import timezone

# def asset_report_view(request):
#     # Get filter parameters
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     min_price = request.GET.get('min_price')
#     max_price = request.GET.get('max_price')
#     asset_type = request.GET.get('asset_type')
#     branch = request.GET.get('branch')
#     department = request.GET.get('department')

#     # Base queryset
#     assets = AssetData.objects.select_related(
#         'asset_type',
#         'branch',
#         'department'
#     ).prefetch_related(
#         'custom_fields',
#         'depreciations'
#     ).all()
    
#     custom_field_names = CustomField.objects.values_list(
#         'field_name', flat=True).distinct()

#     # Apply filters
#     if start_date:
#         assets = assets.filter(purchase_date__gte=start_date)
#     if end_date:
#         assets = assets.filter(purchase_date__lte=end_date)
#     if min_price:
#         assets = assets.filter(price__gte=min_price)
#     if max_price:
#         assets = assets.filter(price__lte=max_price)
#     if asset_type:
#         assets = assets.filter(asset_type_id=asset_type)
#     if branch:
#         assets = assets.filter(branch_id=branch)
#     if department:
#         assets = assets.filter(department_id=department)
        
        
#     assets_data = []
#     for asset in assets:
#         asset_dict = {
#             'asset_code': asset.asset_code,
#             'asset_name': asset.asset_name,
#             'asset_type': asset.asset_type.name,
#             'model': asset.model,
#             'serial_number': asset.serial_number,
#             'office': asset.office.name,
#             'branch': asset.branch.branch_name,
#             'department': asset.department.dpt_name,
#             'purchase_date': asset.purchase_date,
#             'warranty_info': asset.warranty_info,
#             'price': asset.price,
#             'status': 'Sold' if asset.is_sold else 'Active',
#         }

#     # Calculate summaries
#     total_assets = assets.count()
#     total_value = assets.aggregate(Sum('price'))['price__sum'] or 0
#     active_assets = assets.filter(is_sold=False).count()

#     # Calculate depreciated value
#     total_depreciated_value = 0
#     for asset in assets:
#         depreciation = asset.depreciations.first()  # Get the latest depreciation
#         if depreciation:
#             # Calculate depreciated value based on your business logic
#             # This is a simplified example
#             years_owned = (timezone.now().date() - asset.purchase_date).days / 365
#             if years_owned > depreciation.useful_life:
#                 depreciated_value = depreciation.salvage_value
#             else:
#                 annual_depreciation = (depreciation.purchase_value - depreciation.salvage_value) / depreciation.useful_life
#                 depreciated_value = depreciation.purchase_value - (annual_depreciation * years_owned)
#             total_depreciated_value += depreciated_value

#     # Prepare chart data
#     asset_type_data = list(assets.values('asset_type__name').annotate(
#         count=Count('id')
#     ))
    
#     department_data = list(assets.values('department__dpt_name').annotate(
#         total_value=Sum('price')
#     ))

#     context = {
#         'assets': assets,
#         'total_assets': total_assets,
#         'total_value': total_value,
#         'total_depreciated_value': total_depreciated_value,
#         'active_assets': active_assets,
#         'asset_types': AssetType.objects.all(),
#         'branches': Branches.objects.all(),
#         'departments': Departments.objects.all(),
        
#         # Chart data
#         'asset_type_labels': json.dumps([item['asset_type__name'] for item in asset_type_data]),
#         'asset_type_data': json.dumps([item['count'] for item in asset_type_data]),
#         'department_labels': json.dumps([item['department__dpt_name'] for item in department_data]),
#         'department_values': json.dumps([float(item['total_value'] or 0) for item in department_data])
#     }

#     return render(request, 'asset/asset_report.html', context)
from django.db.models import Sum, Count
from django.utils import timezone
import json
from datetime import datetime
@login_required()
def asset_report_view(request):
    # Get filter parameters and handle type conversion
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    asset_type = request.GET.get('asset_type')
    branch = request.GET.get('branch')
    department = request.GET.get('department')

    # Base queryset
    assets = AssetData.objects.select_related(
        'asset_type',
        'branch',
        'department'
    ).prefetch_related(
        'custom_fields',
        'depreciations'
    ).all()
    
    # Apply filters
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            assets = assets.filter(purchase_date__gte=start_date)
        except ValueError:
            pass

    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            assets = assets.filter(purchase_date__lte=end_date)
        except ValueError:
            pass

    if min_price:
        try:
            min_price = float(min_price)
            assets = assets.filter(price__gte=min_price)
        except ValueError:
            pass

    if max_price:
        try:
            max_price = float(max_price)
            assets = assets.filter(price__lte=max_price)
        except ValueError:
            pass

    if asset_type:
        assets = assets.filter(asset_type_id=asset_type)
    if branch:
        assets = assets.filter(branch_id=branch)
    if department:
        assets = assets.filter(department_id=department)

    # Calculate summaries
    total_assets = assets.count()
    total_value = assets.aggregate(Sum('price'))['price__sum'] or 0
    UZS_TO_USD_RATE = Decimal('0.000082')  # Example fixed rate
    total_value_usd = round(total_value * UZS_TO_USD_RATE, 2)
    active_assets = assets.filter(is_sold=False).count()

    # Calculate depreciation
    total_depreciation = 0
    for asset in assets:
        depreciation = asset.depreciations.first()
        if depreciation:
            years_owned = (timezone.now().date() - asset.purchase_date).days / 365
            if years_owned > depreciation.useful_life:
                depreciated_value = depreciation.salvage_value
            else:
                annual_depreciation = (depreciation.purchase_value - depreciation.salvage_value) / depreciation.useful_life
                depreciated_value = depreciation.purchase_value - (annual_depreciation * years_owned)
            total_depreciation += depreciated_value
    
    page = request.GET.get('page', 1)  # Get the current page from the request
    paginator = Paginator(assets, 10)  # Show 10 assets per page

    try:
        assets_paginated = paginator.page(page)
    except PageNotAnInteger:
        assets_paginated = paginator.page(1)
    except EmptyPage:
        assets_paginated = paginator.page(paginator.num_pages)

    context = {
        'assets': assets_paginated,
        'total_assets': total_assets,
        'total_value': total_value,
        'total_value_usd':total_value_usd,
        'total_depreciation': total_depreciation,
        'total_active_assets': active_assets,
        'asset_types': AssetType.objects.all(),
        'branches': Branches.objects.all(),
        'departments': Departments.objects.all(),
    }

    return render(request, 'asset/asset_report.html', context)

def calculate_total_depreciation(assets):
    total_depreciated_value = 0
    for asset in assets:
        depreciation = asset.depreciations.first()
        if depreciation:
            years_owned = (datetime.now().date() - asset.purchase_date).days / 365
            depreciation_schedule = depreciation.calculate_depreciation()
            current_value = float(asset.price)
            for year, amount in depreciation_schedule:
                if year <= years_owned:
                    current_value -= amount
            total_depreciated_value += max(current_value, 0)
    return total_depreciated_value

def get_monthly_acquisitions(assets):
    return assets.annotate(
        month=TruncMonth('purchase_date')
    ).values('month').annotate(
        count=Count('id'),
        value=Sum('price')
    ).order_by('month')
    
    
from django.http import JsonResponse
from .utils import get_monthly_acquisitions, get_asset_statistics, calculate_depreciation_summary

def get_chart_data(request):
    """API endpoint for chart data"""
    assets = AssetData.objects.all()
    
    # Apply filters from request.GET if needed
    # ... (filtering code as in main view)

    # Get monthly acquisition data
    monthly_data = get_monthly_acquisitions(assets)
    
    # Get asset type distribution
    type_distribution = assets.values('asset_type__name').annotate(
        count=Count('id'),
        value=Sum('price')
    )
    
    # Get department distribution
    dept_distribution = assets.values('department__dpt_name').annotate(
        count=Count('id'),
        value=Sum('price')
    )
    
    return JsonResponse({
        'monthly_data': list(monthly_data),
        'type_distribution': list(type_distribution),
        'dept_distribution': list(dept_distribution)
    })




from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
import json
import logging
from .models import AssetData, CustomField
from authentication.models import  CommissionMembers, ChairMan
from datetime import datetime


def export_excel_report(export_data, user, additional_data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from django.http import HttpResponse
    from datetime import datetime
    import os
    from django.conf import settings

    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Report"

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Header: Add company logo if available
    head_office = HeadOffice.objects.first()
    if head_office and head_office.logo_image:
        logo_path = os.path.join(settings.MEDIA_ROOT, str(head_office.logo_image))
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 100
            img.height = 50
            ws.add_image(img, 'A1')
            ws.merge_cells('A1:A3')

    # Add company address below the logo
    if head_office:
        name = f"{head_office.name}\n"
        address = f"{head_office.address}\n"
        phone = f"{head_office.contact_number}\n"
        ws['C1'] = name
        ws['C1'].font = Font(size=10)
        ws['C2'] = address
        ws['C2'].font = Font(size=10)
        ws['C3'] = phone
        ws['C3'].font = Font(size=10)

    # Add report title and date
    ws['A4'] = f"Asset Report - Generated on {datetime.now().strftime('%Y-%m-%d')}"
    ws['A4'].font = Font(bold=True, size=14)
    ws.merge_cells('A4:F4')

    current_row = 8

    # Add asset data with formatted headers
    for row_index, row_data in enumerate(export_data):
        for col_index, value in enumerate(row_data, 1):
            # Format headers by replacing underscores with spaces and capitalizing words
            if row_index == 0:  # Headers
                formatted_value = value.replace('_', ' ').title()
                cell = ws.cell(row=current_row + row_index, column=col_index, value=formatted_value)
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell = ws.cell(row=current_row + row_index, column=col_index, value=value)

    # Move to the row after the table
    current_row = ws.max_row + 3

    # Add signature section header
    ws[f'A{current_row}'] = "Signatures:"
    ws[f'A{current_row}'].font = header_font
    current_row += 2

    # Add commission members signatures on the left side
    if additional_data.get('commission_members'):
        start_col = 1
        for member in additional_data['commission_members']:
            ws.cell(row=current_row, column=start_col, value="Member's Signature")
            ws.cell(row=current_row + 1, column=start_col, value="_________________")
            ws.cell(row=current_row + 2, column=start_col, value=member['name'])
            start_col += 2

    # Add chairman's signature on the right side
    if additional_data.get('chairman'):
        right_col = ws.max_column if ws.max_column > 6 else 6
        
        chairman = additional_data['chairman']
        ws.cell(row=current_row - 1, column=right_col - 1, value="Chairman Details")
        ws.cell(row=current_row - 1, column=right_col - 1).font = header_font
        
        ws.cell(row=current_row, column=right_col - 1, value="Chairman's Signature")
        ws.cell(row=current_row + 1, column=right_col - 1, value="_________________")
        ws.cell(row=current_row + 2, column=right_col - 1, value=chairman['name'])

    # Footer
    current_row += 4
    num_columns_to_merge = 6
    footer_note = "(This document is computer generated and does not require the Registrar's signature or the Company's stamp to be considered valid.)"
    ws[f'A{current_row}'] = footer_note
    ws[f'A{current_row}'].font = Font(italic=True, size=10)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_columns_to_merge)

    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                cell_value = str(cell.value) if cell.value is not None else ''
                max_length = max(max_length, len(cell_value))
            except:
                continue
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asset_report.xlsx'
    wb.save(response)
    return response




def format_date(value):
    """
    Helper function to format date fields properly in DD-MM-YYYY format.
    Handles both datetime objects and string dates.
    """
    try:
        # If value is already a datetime object
        if isinstance(value, datetime):
            return value.strftime('%d-%m-%Y')
        
        # If value is a string, try to parse it
        elif isinstance(value, str) and value.strip():
            # Try different possible date formats
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y', '%d/%m/%Y']:
                try:
                    parsed_date = datetime.strptime(value.strip(), fmt)
                    return parsed_date.strftime('%d-%m-%Y')
                except ValueError:
                    continue
            
        # If no valid date format is found, return original value
        return value
    except:
        return value
    
from datetime import datetime, date


def export_pdf_report(export_data, user, additional_data):
    """
    Generate PDF report for asset data
    
    Args:
        export_data: List of lists containing asset data
        user: User object of person generating report
        additional_data: Dict containing extra report data like signatures
    """
    buffer = None
    try:
        # Validate input data
        MAX_ROWS = 5000
        if len(export_data) > MAX_ROWS:
            return JsonResponse({
                'status': 'error',
                'message': f'Export limited to {MAX_ROWS} rows. Please refine your selection.'
            })

        # Create buffer for PDF
        buffer = BytesIO()

        # Get HeadOffice details
        head_office = HeadOffice.objects.first()
        if not head_office:
            raise ValueError("Head office details not configured")

        # Create styles
        styles = getSampleStyleSheet()
        
        # Define all styles
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,  # Center alignment
        )
        
        address_style = ParagraphStyle(
            'AddressStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1  # Center alignment
        )

        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,  # Center alignment
            textColor=colors.grey
        )

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            spaceBefore=50,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#000000'),
        )

        subtitle_style = ParagraphStyle(
            'SubTitle',
            parent=styles['Heading2'],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=10,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#000000')
        )

        normal_style = ParagraphStyle(
            'NormalText',
            parent=styles['Normal'],
            fontSize=10,
            spaceBefore=2,
            spaceAfter=2,
            wordWrap='CJK'
        )

        italic_style = ParagraphStyle(
            'ItalicStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,  # Center alignment
            textColor=colors.black,
            italic=True,
            spaceBefore=20
        )

        def format_date(value):
            """Format date values consistently"""
            try:
                if isinstance(value, (datetime, date)):
                    return value.strftime('%Y-%m-%d')
                return str(value)
            except Exception as e:
                logger.error(f"Date formatting error: {str(e)}")
                return str(value)

        def split_table_data(export_data, columns_per_row=10):
            """Process and validate table data"""
            try:
                if not export_data or not isinstance(export_data, (list, tuple)):
                    raise ValueError("Invalid export data format")
                    
                if not export_data[0]:
                    raise ValueError("No header data available")
                
                total_columns = len(export_data[0])
                if total_columns > 10:
                    raise ValueError(
                        'The number of columns exceeds the maximum limit of 10 columns for PDF export.\n\n'
                        'Please either:\n'
                        '1. Select fewer fields for the PDF export\n'
                        '2. Use Excel export for viewing all columns\n\n'
                        'PDF has limited space and too many columns may affect readability.'
                    )

                # Process headers with validation
                headers = []
                for header in export_data[0]:
                    if header:
                        headers.append(str(header).replace('_', ' ').title()[:30])
                    else:
                        headers.append('N/A')

                # Create table with serial numbers
                table_data = [['S.No'] + headers]
                
                for idx, row in enumerate(export_data[1:], 1):
                    processed_row = [str(idx)]
                    for value in row:
                        try:
                            if isinstance(value, (datetime, date)):
                                formatted_value = format_date(value)
                            elif isinstance(value, str) and any(x in value.lower() for x in ['date', 'time']):
                                # Special handling for date fields
                                formatted_value = format_date(value)
                            else:
                                formatted_value = str(value)
                                if len(formatted_value) > 50:
                                    formatted_value = formatted_value[:47] + "..."
                            processed_row.append(formatted_value)
                        except Exception as e:
                            processed_row.append('Error')
                            logger.error(f"Error processing cell value: {str(e)}")
                    
                    table_data.append(processed_row)

                return table_data

            except Exception as e:
                logger.error(f"Error in split_table_data: {str(e)}")
                raise

        def create_table_with_style(table_data):
            """Create a formatted table with dynamic column widths"""
            try:
                available_width = 650
                sno_width = 35
                
                col_count = len(table_data[0]) - 1
                
                col_lengths = []
                for col in range(col_count):
                    max_length = max(len(str(row[col+1])) for row in table_data)
                    col_lengths.append(max_length)
                
                total_chars = sum(col_lengths)
                remaining_width = available_width - sno_width
                
                col_widths = [sno_width]
                for length in col_lengths:
                    width = (length / total_chars) * remaining_width
                    width = max(width, 50)
                    col_widths.append(width)

                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('WORDWRAP', (0, 0), (-1, -1), True),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                ])

                formatted_data = [
                    [Paragraph(str(cell), normal_style) for cell in row]
                    for row in table_data
                ]

                table = Table(formatted_data, colWidths=col_widths)
                table.setStyle(table_style)
                return table

            except Exception as e:
                logger.error(f"Error creating table: {str(e)}")
                raise

        class PdfReport(SimpleDocTemplate):
            def __init__(self, *args, **kwargs):
                self.head_office = kwargs.pop('head_office')
                self.generated_by = kwargs.pop('generated_by', 'Unknown User')
                self.generated_on = kwargs.pop('generated_on', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                self.header_style = header_style
                self.footer_style = footer_style
                self.address_style = address_style
                super().__init__(*args, **kwargs)

            def header(self, canvas, doc, first_page=False):
                canvas.saveState()
                top_margin = doc.height + doc.topMargin

                if first_page:
                    # Draw logo only on first page
                    logo_y_position = top_margin - 15
                    if self.head_office.logo_image:
                        try:
                            logo_path = os.path.join(settings.MEDIA_ROOT, str(self.head_office.logo_image))
                            if os.path.exists(logo_path):
                                canvas.drawImage(logo_path, doc.leftMargin, logo_y_position, 
                                                  width=2*inch, height=0.8*inch, preserveAspectRatio=True)
                        except Exception as e:
                            logger.error(f"Error drawing logo: {str(e)}")

                    # Draw company details only on first page
                    header_text_y_position = logo_y_position - 15
                    address_text = f"""
                    <para alignment="right">
                    <b>{self.head_office.name}</b><br/>
                    {self.head_office.address}<br/>
                    Contact: {self.head_office.contact_number}
                    </para>
                    """
                    p_address = Paragraph(address_text, self.address_style)
                    p_address.wrapOn(canvas, doc.width/2, doc.topMargin)
                    p_address.drawOn(canvas, doc.width/2 + doc.leftMargin, header_text_y_position)

                    # Draw separator line
                    canvas.line(doc.leftMargin, top_margin - 55,
                                doc.width + doc.leftMargin, top_margin - 55)

                # Draw generation info on all pages
                generated_info_y_position = top_margin - (70 if first_page else 20)
                generated_info_left = f"""
                <para alignment="left">
                Report Generated on: {self.generated_on}
                </para>
                """
                p_info_left = Paragraph(generated_info_left, self.header_style)
                p_info_left.wrapOn(canvas, doc.width/2 - doc.leftMargin, doc.topMargin)
                p_info_left.drawOn(canvas, doc.leftMargin, generated_info_y_position)

                generated_info_right = f"""
                <para alignment="right">
                Report Generated by: {self.generated_by}
                </para>
                """
                p_info_right = Paragraph(generated_info_right, self.header_style)
                p_info_right.wrapOn(canvas, doc.width/2 - doc.leftMargin, doc.topMargin)
                p_info_right.drawOn(canvas, doc.width/2 + doc.leftMargin, generated_info_y_position)

                canvas.restoreState()

            def footer(self, canvas, doc):
                canvas.saveState()
                footer_margin = doc.bottomMargin - 10

                canvas.line(doc.leftMargin, footer_margin + 15,
                            doc.width + doc.leftMargin, footer_margin + 15)

                footer_text = "@ Gaude AI Solution - Powered by Asset Trace"
                p = Paragraph(footer_text, self.footer_style)
                p.wrapOn(canvas, doc.width, doc.bottomMargin)
                p.drawOn(canvas, doc.leftMargin, footer_margin)

                page_num = canvas.getPageNumber()
                canvas.drawRightString(doc.width + doc.leftMargin, footer_margin,
                                        f"Page {page_num}")

                canvas.restoreState()

        def myFirstPage(canvas, doc):
            doc.header(canvas, doc, first_page=True)
            doc.footer(canvas, doc)

        def myLaterPages(canvas, doc):
            doc.header(canvas, doc, first_page=False)
            doc.footer(canvas, doc)

        # Force garbage collection
        if len(export_data) > 1000:
            gc.collect()

        # Build PDF Content
        elements = []

        # Add title section with extra spacing
        elements.append(Spacer(1, 40))
        elements.append(Paragraph('Assets Report', title_style))
        elements.append(Spacer(1, 20))

        # Process and add table
        table_data = split_table_data(export_data)
        table = create_table_with_style(table_data)
        elements.append(table)
        elements.append(Spacer(1, 20))

       # Add signature section if needed
        # Add signature section if needed
        if additional_data.get('commission_members') or additional_data.get('chairman'):
            details_data = []
            left_content = []
            right_content = []

            if additional_data.get('commission_members'):
                left_content.append(Paragraph("Commission Members", subtitle_style))
                for member in additional_data['commission_members']:
                    left_content.append(Paragraph(member['name'], normal_style))
                    left_content.append(Spacer(1, 5))

            if additional_data.get('chairman'):
                
                chairman = additional_data['chairman']
                right_content.append(Paragraph(f"Name: {chairman['name']}<br/>_________________", normal_style))
                if 'email' in chairman:
                    right_content.append(Paragraph(f"Email: {chairman['email']}", normal_style))
                right_content.append(Spacer(1, 5))

            max_length = max(len(left_content), len(right_content))
            while len(left_content) < max_length:
                left_content.append(Paragraph("", normal_style))
            while len(right_content) < max_length:
                right_content.append(Paragraph("", normal_style))

            for i in range(max_length):
                details_data.append([left_content[i], right_content[i]])

            details_table = Table(details_data, colWidths=['50%', '50%'])
            details_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 20),
                ('RIGHTPADDING', (0, 0), (-1, -1), 20),
            ]))
            elements.append(details_table)
            elements.append(Spacer(1, 30))

            elements.append(Paragraph("Signatures", subtitle_style))
            elements.append(Spacer(1, 20))

            signature_data = []
            signature_row = []

            if additional_data.get('commission_members'):
                for member in additional_data['commission_members']:
                    signature_row.append(
                        Paragraph(f"{member['name']}<br/>_________________", normal_style)
                    )

            

            signature_data.append(signature_row)

            num_signatures = len(signature_row)
            col_width = 540 / num_signatures

            signature_table = Table(signature_data, colWidths=[col_width] * num_signatures)
            signature_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(signature_table)



        # Add footer note
        elements.append(Spacer(1, 30))
        footer_note = Paragraph(
            "(This document is computer generated and does not require the Registrar's signature or the Company's stamp in order to be considered valid.)",
            italic_style
        )
        elements.append(footer_note)

        # Generate PDF
        doc = PdfReport(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=72,
            bottomMargin=36,
            head_office=head_office,
            generated_by=user.username,
            generated_on=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )

        # Build the PDF
        doc.build(elements, onFirstPage=myFirstPage, onLaterPages=myLaterPages)

        # Create the HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="asset_report.pdf"'
        response['X-Content-Type-Options'] = 'nosniff'

        # Get PDF from buffer and write to response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred while generating the PDF: {str(e)}'
        })

@login_required()
def export_report(request):
    if request.method == 'POST':
        try:
            user = request.user
            data = json.loads(request.body)
            
            # Validate request data
            if not data:
                return JsonResponse({'error': 'No data provided'}, status=400)

            selected_fields = data.get('fields', [])
            if not selected_fields:
                return JsonResponse({'error': 'No fields selected'}, status=400)

            selected_assets = data.get('assets', [])
            if not selected_assets:
                return JsonResponse({'error': 'No assets selected'}, status=400)

            report_type = data.get('reportType', 'generic')
            export_format = data.get('format')
            
            if export_format not in ['excel', 'pdf']:
                return JsonResponse({'error': 'Invalid export format'}, status=400)
            
            

            # Get additional data based on report type
            additional_data = {
                'commission_members': [],
                'chairman': None
            }
            
            if report_type in ['commission', 'full']:
                commission_members = CommissionMembers.objects.all()
                additional_data['commission_members'] = [
                    {
                        'name': member.name,
                        
                    } for member in commission_members
                ]
            
            if report_type in ['chairman', 'full']:
                chairman = ChairMan.objects.first()
                if chairman:
                    additional_data['chairman'] = {
                        'name': chairman.name,
                        
                    }

            # Get assets with selected IDs
            assets_query = AssetData.objects.select_related(
                'asset_type',
                'branch',
                'department'
            ).prefetch_related(
                'custom_fields'
            )
            assets = assets_query.all()
            print('assets:',assets)

            if not assets.exists():
                return JsonResponse({'error': 'No assets found'}, status=404)

            # Get all unique custom fields
            custom_fields = CustomField.objects.filter(
                asset__in=assets
            ).values_list('field_name', flat=True).distinct()

            # Prepare export data
            export_data = []
            include_headers = data.get('includeHeaders', True)

            if include_headers:
                headers = [field for field in selected_fields]
                export_data.append(headers)

            # Add asset data rows
            for asset in assets:
                row_data = []
                for field in selected_fields:
                    try:
                        if field in [custom.field_name for custom in asset.custom_fields.all()]:
                            custom_value = asset.custom_fields.filter(field_name=field).first()
                            row_data.append(custom_value.field_value if custom_value else '')
                        else:
                            value = get_asset_field_value(asset, field)
                            row_data.append(value)
                    except Exception as e:
                        logger.error(f"Error processing field {field} for asset {asset.id}: {str(e)}")
                        row_data.append('')

                export_data.append(row_data)

            # Generate and return appropriate response
            if export_format == 'excel':
                return export_excel_report(export_data, user, additional_data)
            else:  # PDF
                return export_pdf_report(export_data, user, additional_data)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Export error: {str(e)}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Only POST method is allowed'}, status=405)




def get_asset_field_value(asset, field):
    """Helper function to get asset field values"""
    field_mapping = {
        'asset_code': lambda: asset.asset_code or '',
        'name': lambda: asset.asset_name or '',
        'asset_type': lambda: asset.asset_type.name if asset.asset_type else '',
        'branch': lambda: asset.branch.branch_name if asset.branch else '',
        'department': lambda: asset.department.dpt_name if asset.department else '',
        'purchase_date': lambda: asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
        'value': lambda: f"${asset.price:,.2f}" if asset.price is not None else '',
        'status': lambda: 'Sold' if asset.is_sold else 'Active',
        'serial_number': lambda: asset.serial_number or '',
        'warranty_info': lambda: asset.warranty_info or '',
        'custodian': lambda: asset.custodian or '',
        'sold_date': lambda: asset.sold_date.strftime('%Y-%m-%d') if asset.sold_date else '',
    }
    
    return field_mapping.get(field, lambda: '')()



# from openpyxl.drawing.image import Image 

# def export_excel_report(data, username):
#     try:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Asset Report"
        
#         # Get HeadOffice details
#         head_office = HeadOffice.objects.first()
        
#         # Define styles
#         header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
#         header_fill = PatternFill(start_color='4681f4', end_color='4681f4', fill_type='solid')
#         header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
#         cell_font = Font(name='Arial', size=10)
#         cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
#         company_font = Font(name='Arial', size=12, bold=True)
#         address_font = Font(name='Arial', size=10)
#         info_font = Font(name='Arial', size=9, italic=True)
        
#         border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )

#         # Insert rows for company details
#         ws.insert_rows(1, 6)  # Insert 6 rows at the top

#         # Add company logo
        

#         # Add company details
#         ws['D1'] = head_office.name
#         ws['D1'].font = company_font
#         ws['D1'].alignment = Alignment(horizontal='left')
        
#         ws['D2'] = head_office.address
#         ws['D2'].font = address_font
#         ws['D2'].alignment = Alignment(horizontal='left')
        
#         ws['D3'] = f"Contact: {head_office.contact_number}"
#         ws['D3'].font = address_font
#         ws['D3'].alignment = Alignment(horizontal='left')

#         # Add separator line
#         for col in range(1, ws.max_column + 1):
#             ws.cell(row=4, column=col).border = Border(bottom=Side(style='thin'))

#         # Add generation details
#         ws['A5'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
#         ws['A5'].font = info_font
        
#         ws['D5'] = f"Generated by: {username}"
#         ws['D5'].font = info_font
#         ws['D5'].alignment = Alignment(horizontal='right')

#         # Write and style data (starting from row 7)
#         row_offset = 7
#         for row_idx, row in enumerate(data, row_offset):
#             for col_idx, value in enumerate(row, 1):
#                 cell = ws.cell(row=row_idx, column=col_idx, value=value)
#                 cell.border = border
                
#                 if row_idx == row_offset:  # Header row
#                     cell.font = header_font
#                     cell.fill = header_fill
#                     cell.alignment = header_alignment
#                 else:  # Data rows
#                     cell.font = cell_font
#                     cell.alignment = cell_alignment

#         # Adjust column widths
#         for col in ws.columns:
#             max_length = 0
#             col_letter = get_column_letter(col[0].column)
            
#             # Find the maximum length in the column
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
            
#             # Calculate adjusted width (with limits)
#             adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50
#             ws.column_dimensions[col_letter].width = adjusted_width

#         # Set row heights
#         ws.row_dimensions[row_offset].height = 30  # Header row
#         for row in range(row_offset + 1, len(data) + row_offset):
#             ws.row_dimensions[row].height = 25

#         # Freeze panes after company details and header
#         ws.freeze_panes = f'A{row_offset + 1}'

#         # Add footer
        

#         # Create response
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
        
#         # Generate filename with timestamp and username
#         timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#         response['Content-Disposition'] = f'attachment; filename="asset_report_{username}_{timestamp}.xlsx"'
        
#         # Save to response
#         wb.save(response)
#         return response
        
#     except Exception as e:
#         logger.error(f"Excel export error: {str(e)}", exc_info=True)
#         return JsonResponse({'error': f"Excel export failed: {str(e)}"}, status=500)
    
    
    
    

def get_custom_fields(request):
    if request.method == 'GET':
        try:
            # Fetch all custom fields
            custom_fields = CustomField.objects.values('field_name')

            # Use a set to avoid duplicates
            unique_fields = set()
            custom_fields_list = []

            for field in custom_fields:
                field_name = field['field_name']
                if field_name not in unique_fields:
                    unique_fields.add(field_name)
                    custom_fields_list.append({
                        'value': field_name,
                        'label': field_name.title()
                    })

            return JsonResponse(custom_fields_list, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)













# from reportlab.lib.pagesizes import landscape, inch
# from reportlab.lib import colors
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from reportlab.lib.units import inch
# from reportlab.lib.enums import TA_CENTER, TA_RIGHT
# from django.http import HttpResponse
# from datetime import datetime
# import os
# from django.conf import settings

# def export_pdf_report(data, username):
#     """
#     Convert the export data into PDF format and handle large tables by splitting them.
#     Includes Head Office logo and name and adds Asset Code to the table.
#     """
#     response = HttpResponse(content_type='application/pdf')
#     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#     response['Content-Disposition'] = f'attachment; filename="asset_report_{username}_{timestamp}.pdf"'

#     # Get HeadOffice details
#     head_office = HeadOffice.objects.first()

#     # Create custom styles
#     styles = getSampleStyleSheet()
#     title_style = ParagraphStyle(
#         'CustomTitle',
#         parent=styles['Heading1'],
#         fontSize=16,
#         spaceAfter=20,
#         alignment=TA_CENTER
#     )
#     header_style = ParagraphStyle(
#         'HeaderStyle',
#         parent=styles['Normal'],
#         fontSize=9,
#         alignment=TA_RIGHT
#     )

#     table_style = TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#         ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('FONTSIZE', (0, 0), (-1, 0), 8),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
#         ('BACKGROUND', (0, 1), (-1, -1), colors.white),
#         ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
#         ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#         ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
#         ('FONTSIZE', (0, 1), (-1, -1), 8),
#         ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
#         ('WORDWRAP', (0, 0), (-1, -1), True),
#     ])

#     def wrap_text(text, max_length=150):
#         """
#         Helper function to wrap long text in a cell.
#         """
#         if len(str(text)) > max_length:
#             text = str(text)[:max_length] + "..."
#         return text

#     # Define header with logo and name of the company
#     def add_header(canvas, doc):
#         canvas.saveState()

#         # Display the logo (if available)
#         logo_y_position = doc.height + doc.topMargin - 15
#         if head_office.logo_image:
#             logo_path = os.path.join(settings.MEDIA_ROOT, str(head_office.logo_image))
#             canvas.drawImage(logo_path, doc.leftMargin, logo_y_position, width=2*inch, height=0.8*inch, preserveAspectRatio=True)

#         # Head office name and address
#         header_text_y_position = logo_y_position - 15
#         address_text = f"""
#         <b>{head_office.name}</b><br/>
#         {head_office.address}<br/>
#         Contact: {head_office.contact_number}
#         """
#         p_address = Paragraph(address_text, header_style)
#         p_address.wrapOn(canvas, doc.width, doc.topMargin)
#         p_address.drawOn(canvas, doc.leftMargin, header_text_y_position)

#         canvas.restoreState()

#     # Define footer (optional)
#     def add_footer(canvas, doc):
#         canvas.saveState()
#         footer_margin = doc.bottomMargin - 10
#         footer_text = "@ Gaude AI Solution - Powered by Asset Trace"
#         p = Paragraph(footer_text, header_style)
#         p.wrapOn(canvas, doc.width, doc.bottomMargin)
#         p.drawOn(canvas, doc.leftMargin, footer_margin)

#         page_num = canvas.getPageNumber()
#         canvas.drawRightString(doc.width + doc.leftMargin, footer_margin, f"Page {page_num}")

#         canvas.restoreState()

#     elements = []
#     elements.append(Paragraph('Assets Report', title_style))

#     # Add Asset Code to the table headers if it's part of your data
#     max_columns_per_table = 10  # Define how many columns you can fit per table
#     total_columns = len(data[0])

#     # Loop through column ranges and split large tables into sections
#     for col_start in range(0, total_columns, max_columns_per_table):
#         col_end = min(col_start + max_columns_per_table, total_columns)

#         # Prepare the table section with only a subset of columns
#         partial_table_data = []

#         # Add headers (include Asset Code as the first column)
#         partial_table_data.append([wrap_text(cell) for cell in data[0][col_start:col_end]])

#         # Add the rows
#         for row in data[1:]:
#             partial_table_data.append([wrap_text(cell) for cell in row[col_start:col_end]])

#         # Set column widths (adjust depending on the number of columns you want to fit)
#         col_widths = [1.5 * inch] * (col_end - col_start)
#         table = Table(partial_table_data, colWidths=col_widths, repeatRows=1)
#         table.setStyle(table_style)

#         # Add table to document
#         elements.append(table)
#         elements.append(Paragraph("<br/>", styles['Normal']))  # Add a gap between tables

#     # Create PDF document
#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape((12 * inch, 8.5 * inch)),  # Use landscape mode for wide tables
#         rightMargin=10,
#         leftMargin=10,
#         topMargin=90,
#         bottomMargin=30,
#     )

#     # Build the PDF with header and footer callbacks
#     try:
#         doc.build(elements, onFirstPage=add_header, onLaterPages=add_footer)
#         return response
#     except Exception as e:
#         logger.error(f"PDF export error: {str(e)}", exc_info=True)
#         return JsonResponse({'error': f"PDF export failed: {str(e)}"}, status=500)




# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from .models import AssetData, AssetType
from  authentication.models import HeadOffice, Branches, Departments

import pandas as pd
import zipfile
from io import BytesIO
from django.core.files import File
from django.core.files.base import ContentFile
import zipfile
@login_required()
def upload_assets(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            images_zip = request.FILES.get('images_zip')
            bills_zip = request.FILES.get('bills_zip')

            # Read Excel file
            df = pd.read_excel(excel_file)
            
            # Initialize dictionaries for files
            image_files = {}
            bill_files = {}
            
            # Process images ZIP if provided
            if images_zip:
                try:
                    zip_file = zipfile.ZipFile(images_zip)
                    # List all files in the zip
                    for filename in zip_file.namelist():
                        if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                            # Read the file content
                            file_content = zip_file.read(filename)
                            # Store with both full path and just filename
                            image_files[filename] = file_content  # Store with full path
                            # Also store with just the filename (without directory)
                            base_filename = filename.split('/')[-1]
                            image_files[base_filename] = file_content
                    zip_file.close()
                except Exception as e:
                    messages.error(request, f'Error processing images ZIP: {str(e)}')
                    return redirect('upload_assets')

            # Process bills ZIP if provided
            if bills_zip:
                try:
                    zip_file = zipfile.ZipFile(bills_zip)
                    for filename in zip_file.namelist():
                        if filename.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png')):
                            file_content = zip_file.read(filename)
                            bill_files[filename] = file_content
                    zip_file.close()
                except Exception as e:
                    messages.error(request, f'Error processing bills ZIP: {str(e)}')
                    return redirect('upload_assets')

            # Process each row in Excel
            for index, row in df.iterrows():
                try:
                    # Create asset
                    asset_type = AssetType.objects.get(name=row['asset_type'])
                    office = HeadOffice.objects.get(name=row['office'])
                    branch = Branches.objects.get(branch_name=row['branch'])
                    department = Departments.objects.get(dpt_name=row['department'])

                    asset = AssetData.objects.create(
                        asset_type=asset_type,
                        asset_code=row['asset_code'],
                        office=office,
                        branch=branch,
                        department=department,
                        asset_name=row['asset_name'],
                        model=row.get('model'),
                        serial_number=row.get('serial_number'),
                        purchase_date=pd.to_datetime(row.get('purchase_date')).date() if pd.notnull(row.get('purchase_date')) else None,
                        warranty_info=row.get('warranty_info'),
                        price=row['price'],
                        is_sold=row.get('is_sold', False),
                        custodian=row['custodian']
                    )

                    # Process custom fields
                    for col in df.columns:
                        if col.startswith('custom_'):
                            field_name = col.replace('custom_', '')
                            field_value = row[col]
                            if pd.notnull(field_value):
                                CustomField.objects.create(
                                    asset=asset,
                                    field_name=field_name,
                                    field_value=str(field_value)
                                )

                    # Process images
                    if pd.notnull(row.get('images')):
                        # Split the image filenames and clean them
                        image_names = [name.strip() for name in str(row['images']).split(',')]
                        
                        # Debug print
                        print(f"Looking for images: {image_names}")
                        print(f"Available images in ZIP: {list(image_files.keys())}")

                        for image_name in image_names:
                            # Look for exact match and case-insensitive match
                            if image_name in image_files:
                                image_content = image_files[image_name]
                            else:
                                # Try case-insensitive match
                                image_name_lower = image_name.lower()
                                matching_files = [f for f in image_files.keys() 
                                               if f.lower() == image_name_lower]
                                if matching_files:
                                    image_content = image_files[matching_files[0]]
                                else:
                                    print(f"Image not found: {image_name}")
                                    continue

                            try:
                                # Create and save the image
                                asset_image = AssetImage(asset=asset)
                                asset_image.image.save(
                                    image_name,
                                    ContentFile(image_content),
                                    save=True
                                )
                                print(f"Successfully saved image: {image_name}")
                            except Exception as e:
                                print(f"Error saving image {image_name}: {str(e)}")
                                continue

                    # Process bill
                    if pd.notnull(row.get('bill')):
                        bill_name = str(row['bill']).strip()
                        if bill_name in bill_files:
                            bill = AssetBill(asset=asset)
                            bill.bill_file.save(
                                bill_name,
                                ContentFile(bill_files[bill_name]),
                                save=True
                            )

                except Exception as e:
                    messages.error(request, f'Error in row {index + 2}: {str(e)}')
                    continue

            messages.success(request, 'Assets uploaded successfully!')
            return redirect('asset_list')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('upload_assets')

    return render(request, 'asset/upload_assets.html')

# def upload_assets(request):
#     if request.method == 'POST':
#         try:
#             # Check if files are present in request
#             if 'excel_file' not in request.FILES:
#                 messages.error(request, 'Excel file is required!')
#                 return redirect('upload_assets')

#             excel_file = request.FILES['excel_file']
#             images_zip = request.FILES.get('images_zip')
#             bills_zip = request.FILES.get('bills_zip')

#             # Read Excel file
#             df = pd.read_excel(excel_file)
            
#             # Initialize dictionaries for files
#             image_files = {}
#             bill_files = {}
            
#             # Process images ZIP if provided
#             if images_zip:
#                 try:
#                     with zipfile.ZipFile(images_zip) as z:
#                         for filename in z.namelist():
#                             if not filename.endswith('/'):  # Skip directories
#                                 if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
#                                     with z.open(filename) as f:
#                                         image_files[filename] = f.read()
#                 except zipfile.BadZipFile:
#                     messages.error(request, 'Invalid images ZIP file!')
#                     return redirect('upload_assets')

#             # Process bills ZIP if provided
#             if bills_zip:
#                 try:
#                     with zipfile.ZipFile(bills_zip) as z:
#                         for filename in z.namelist():
#                             if not filename.endswith('/'):  # Skip directories
#                                 if filename.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png')):
#                                     with z.open(filename) as f:
#                                         bill_files[filename] = f.read()
#                 except zipfile.BadZipFile:
#                     messages.error(request, 'Invalid bills ZIP file!')
#                     return redirect('upload_assets')

#             # Process each row in Excel
#             for index, row in df.iterrows():
#                 try:
#                     # Create asset
#                     asset_type = AssetType.objects.get(name=row['asset_type'])
#                     office = HeadOffice.objects.get(name=row['office'])
#                     branch = Branches.objects.get(branch_name=row['branch'])
#                     department = Departments.objects.get(dpt_name=row['department'])

#                     asset = AssetData.objects.create(
#                         asset_type=asset_type,
#                         asset_code=row['asset_code'],
#                         office=office,
#                         branch=branch,
#                         department=department,
#                         asset_name=row['asset_name'],
#                         model=row.get('model'),
#                         serial_number=row.get('serial_number'),
#                         purchase_date=pd.to_datetime(row.get('purchase_date')).date() if pd.notnull(row.get('purchase_date')) else None,
#                         warranty_info=row.get('warranty_info'),
#                         price=row['price'],
#                         is_sold=row.get('is_sold', False)
#                     )

#                     # Process custom fields
#                     for col in df.columns:
#                         if col.startswith('custom_'):
#                             field_name = col.replace('custom_', '')
#                             field_value = row[col]
#                             if pd.notnull(field_value):
#                                 CustomField.objects.create(
#                                     asset=asset,
#                                     field_name=field_name,
#                                     field_value=str(field_value)
#                                 )

#                     # Process images if available in excel and zip
#                     if pd.notnull(row.get('images')):
#                         image_list = [img.strip() for img in str(row['images']).split(',')]
#                         for image_name in image_list:
#                             if image_name in image_files:
#                                 image = AssetImage(asset=asset)
#                                 image.image.save(
#                                     image_name,
#                                     ContentFile(image_files[image_name]),
#                                     save=True
#                                 )

#                     # Process bill if available in excel and zip
#                     if pd.notnull(row.get('bill')):
#                         bill_name = str(row['bill']).strip()
#                         if bill_name in bill_files:
#                             bill = AssetBill(asset=asset)
#                             bill.bill_file.save(
#                                 bill_name,
#                                 ContentFile(bill_files[bill_name]),
#                                 save=True
#                             )

#                 except Exception as e:
#                     messages.error(request, f'Error in row {index + 2}: {str(e)}')
#                     continue

#             messages.success(request, 'Assets uploaded successfully!')
#             return redirect('asset_list')

#         except Exception as e:
#             messages.error(request, f'Error: {str(e)}')
#             return redirect('upload_assets')

#     return render(request, 'asset/upload_assets.html')
@login_required()
def download_template(request):
    data = {
        'asset_type': [
            'Computers'
        ],
        'asset_code': [
            'COMP001',
            
        ],
        'office': [
            'TECHNOCORP IT Services and IT Consulting',
            
        ],
        'branch': [
            'TECHNOCORP IT Services and IT Consulting',
            
        ],
        'department': [
            'IT',
            
        ],
        'asset_name': [
            'Dell Laptop',
            
        ],
        'model': [
            'Latitude 5420',
            
        ],
        'serial_number': [
            'SN123456',
            
        ],
        'purchase_date': [
            '2024-01-01',
            
        ],
        'warranty_info': [
            '2 years',
            
        ],
        'price': [
            1500.00,
           
        ],
        'is_sold': [
            False,
           
        ],
        'custodian': [
            "name",
           
        ],
        # Custom Fields (prefix with custom_)
        'custom_manufacturer': [
            'Dell',
            
        ],
        'custom_location': [
            'Floor 2, Room 201',
            
        ],
        'custom_condition': [
            'New',
            
        ],
        'custom_assigned_to': [
            'John Doe',
            
        ],
        'custom_maintenance_schedule': [
            'Every 6 months',
            
        ],
        # Images (comma-separated list of filenames)
        'images': [
            'COMP001_front.jpg, COMP001_back.jpg, COMP001_side.jpg',
            
        ],
        # Bills (single file per asset)
        'bill': [
            'COMP001_bill.pdf',
            
        ],
        # Additional custom fields for different asset types
        'custom_processor': [
            'Intel i7 11th Gen',
           
            
        ],
        'custom_ram': [
            '16GB',
           
        ],
        'custom_storage': [
            '512GB SSD',
           
        ],
        'custom_dimensions': [
            
            '160x80x75 cm',
            
        ]
    }
    
    df = pd.DataFrame(data)
    
    # Create Excel writer with xlsxwriter engine
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="asset_template.xlsx"'
    
    # Write to Excel with formatting
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Template', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Template']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        
        # Format the header row
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
        # Set column widths
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.set_column(idx, idx, max_length + 2)
            
        # Add a new sheet with instructions
        instructions = workbook.add_worksheet('Instructions')
        instructions.write('A1', 'Asset Upload Instructions:', workbook.add_format({'bold': True, 'font_size': 12}))
        instructions_text = [
            'Required Fields:',
            '- asset_type: Must match exactly with existing asset types in the system',
            '- asset_code: Unique identifier for each asset',
            '- office: Must match existing office names',
            '- branch: Must match existing branch names',
            '- department: Must match existing department names',
            '- asset_name: Name/description of the asset',
            '- price: Numeric value',
            '',
            'Images:',
            '- List image filenames separated by commas',
            '- Images must be included in the ZIP file',
            '- Supported formats: .jpg, .jpeg, .png',
            '',
            'Bills:',
            '- Single bill file per asset',
            '- Must be included in the bills ZIP file',
            '- Supported formats: .pdf, .jpg, .jpeg, .png',
            '',
            'Custom Fields:',
            '- All fields starting with "custom_" will be saved as custom fields',
            '- Leave empty if not applicable',
            '',
            'File Naming Convention:',
            '- Images: ASSETCODE_description.extension (e.g., COMP001_front.jpg)',
            '- Bills: ASSETCODE_bill.extension (e.g., COMP001_bill.pdf)'
        ]
        
        for i, text in enumerate(instructions_text):
            instructions.write(i + 2, 0, text)
        
        instructions.set_column(0, 0, 70)
    
    return response